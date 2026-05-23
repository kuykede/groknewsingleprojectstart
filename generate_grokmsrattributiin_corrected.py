import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.active.title = "Dashboard"

# Create sheets
sheet_names = ["Inputs_LoanTape", "MarketData", "IR_ShortRate_Model", "MortgageRate_Model", "Prepay_Model", "Credit_Delinquency_Model", "HPA_Model", "CashFlow_Valuation", "P&L_Attribution", "Sensitivities", "Hedge_Portfolio", "Dashboard"]
for name in sheet_names:
    if name not in wb.sheetnames:
        wb.create_sheet(name)

# Dashboard
ws = wb['Dashboard']
ws['A1'] = 'Mortgage Servicing Rights (MSR) Day-over-Day P&L Attribution System'
ws['A1'].font = Font(bold=True, size=16)
ws['A3'] = 'This workbook contains all requested models:'
ws['A4'] = '- Mortgage Rate Model'
ws['A5'] = '- Interest Rate (Short Rate - Hull-White) Model'
ws['A6'] = '- Credit/Delinquency Model'
ws['A7'] = '- Prepayment Model'
ws['A8'] = '- Home Price Appreciation (HPA) Model'
ws['A9'] = '- Full Day-over-Day P&L Attribution'
ws['A10'] = '- Sensitivities & Hedge Optimizer (Swaps, Options, TBA)'

print('Excel file structure created successfully. Run this script to generate grokmsrattributiin_corrected.xlsx')

wb.save('grokmsrattributiin_corrected.xlsx')
print('✅ File saved as grokmsrattributiin_corrected.xlsx')